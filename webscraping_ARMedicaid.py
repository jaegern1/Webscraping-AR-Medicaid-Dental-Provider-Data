from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchFrameException
import time
import openpyxl

def extract_data(driver):
    #drill down to rows
    div = driver.find_element(By.ID,"dnn_ctr604_SearchProvider_ResultsPanel")
    table = div.find_element(By.ID,"dnn_ctr604_SearchProvider_ProviderSearchDataGrid")
    tbody = table.find_element(By.TAG_NAME,"tbody")
    rows = tbody.find_elements(By.TAG_NAME,"tr")

    #write info to excel
    for row in rows:
        cols = row.find_elements(By.XPATH,"./td")
        if len(cols) == 0: continue

        data = []
        data.append(cols[0].text.strip())
        data.append(cols[1].text.strip())
        data.append(cols[2].text.strip())
        data.append(cols[3].text.strip())

        print(data)

        appendxlsx("dentalgroup.xlsx",data)

def load_zipcodes(name):
    zipwb = openpyxl.load_workbook(name) #load zipcodes.xlsx
    zipsheet = zipwb.active #select sheet

    num_rows = zipsheet.max_row
    zipcodes = []
    
    #adds zipcodes from .xlsx to list
    for x in range(1, num_rows+1):
        zipcodes.append(zipsheet.cell(row=x, column=1).value) #first column of zipcodes.xlsx

    return zipcodes

def load_counties(name):
    countieswb = openpyxl.load_workbook(name) #loads zipcodes.xlsx
    countiessheet = countieswb.active #selects sheet

    num_rows = countiessheet.max_row
    counties = []
    
    #adds counties from .xlsx to list
    for x in range(1, num_rows+1):
        counties.append(countiessheet.cell(row=x, column=4).value) #last column of .xlsx

    return counties
def collect_data(zipcode,county,type):
    #load webpage
    driver.get("https://portal.mmis.arkansas.gov/armedicaid/member/Resources/SearchProviders/tabid/97/Default.aspx")
    time.sleep(5)

    #finds query elements
    zipbox = driver.find_element(By.NAME, "dnn$ctr604$SearchProvider$ZipCodeCmnTextBox$Control")
    search = driver.find_element(By.NAME, "dnn$ctr604$SearchProvider$SearchProviderCmnButton")
    provider = Select(driver.find_element(By.NAME, "dnn$ctr604$SearchProvider$ProviderTypeCmnDropDownList$Control"))
    results = Select(driver.find_element(By.NAME, "dnn$ctr604$SearchProvider$ResultPerPageCmnDropDownList$Control"))

    #executes query
    zipbox.clear()
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE)
    zipbox.send_keys(Keys.BACKSPACE) #search box can autopopulate with multiple spaces
    zipbox.send_keys(zipcode)
    time.sleep(5)
    provider.select_by_visible_text(type)
    results.select_by_visible_text('100 per page')
    results.select_by_visible_text('100 per page')
    search.click()

    time.sleep(20)

    #addresses server lag by forever trying to find specific element until it finds it or the no data can be found for zipcodes without available data
    try:
        driver.find_element(By.ID,"dnn_ctr604_SearchProvider_ProviderSearchDataGrid")
    except NoSuchElementException:
        try:
            driver.find_element(By.XPATH,"//*[@id='dnn_ctr604_SearchProvider_ResultsPanel']/div[1]/div/div/span/strong")
            print("No rows for " + zipcode)
            return
        except NoSuchElementException:
            collect_data(zipcode,county,type)

    #drill down to rows
    div = driver.find_element(By.ID,"dnn_ctr604_SearchProvider_ResultsPanel")
    table = div.find_element(By.ID,"dnn_ctr604_SearchProvider_ProviderSearchDataGrid")
    tbody = table.find_element(By.TAG_NAME,"tbody")
    rows = tbody.find_elements(By.TAG_NAME,"tr")

    #write info to excel
    for row in rows:
        cols = row.find_elements(By.XPATH,"./td")
        if len(cols) == 0: continue

        data = []
        data.append(cols[0].text.strip())
        data.append(cols[1].text.strip())
        data.append(cols[2].text.strip())
        data.append(cols[3].text.strip())
        data.append(county)
        print(data)

        appendxlsx("dentaldata.xlsx",data)


    print("done")

#takes row of data from collect_data and writes to file
def appendxlsx(name,row): 
    xlsx = openpyxl.load_workbook(name)
    sheet = xlsx.active
    sheet.append(row)
    xlsx.save(name)
    xlsx.close()

driver = webdriver.Chrome()

zipcodes = load_zipcodes("zipcodes.xlsx")
counties = load_counties("zipcodes.xlsx")

#runs webscraping for all dental provider offices and individual dentists
for x in range(len(zipcodes)):
    print(x+1,zipcodes[x])
    collect_data(zipcodes[x],counties[x],'DENTAL GROUP')
for x in range(len(zipcodes)):
    print(x+1,zipcodes[x])
    collect_data(zipcodes[x],counties[x],'DENTAL')

driver.quit()
